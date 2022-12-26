import openpyxl
wb = openpyxl.load_workbook(filename='address/add.xlsx')
ws = wb.worksheets[0]
ws['A1'].value = 1


for i in range (2,478):
    with open(f"address/text/{i}.txt", 'r', encoding="utf-8")as f:
        add = f.readline()
        print(add)
        ws[f'A{i}'].value = add
        

wb.save('address/add.xlsx')