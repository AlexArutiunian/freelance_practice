import openpyxl

wb = openpyxl.load_workbook(filename='inn.xlsx')
ws = wb.worksheets[0]
ws['A1'].value = 1

n = 10
for i in range (2,n):
    with open(f'inn/{i}.txt', 'r') as f:
        print(i)
        inn_ = f.readline()
        
        ws[f'A{i}'].value = inn_
        

wb.save('inn.xlsx')